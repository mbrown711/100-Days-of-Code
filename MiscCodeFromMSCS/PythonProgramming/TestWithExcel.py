import mysql.connector
from openpyxl import Workbook

def main():

    # Connect to DB -----------------------------------------------------------
    db = mysql.connector.connect( user='root', password='', host='127.0.0.1')
    cur = db.cursor()

    # Create table ------------------------------------------------------------
    database = 'test_database'
    SQL = 'CREATE DATABASE IF NOT EXISTS ' + database + ';'
    cur.execute(SQL)
    db.commit()

    SQL = 'USE ' + database + ';'
    cur.execute(SQL)

    # Create car data ---------------------------------------------------------
    cars_table_name = 'cars'
    SQL = (
        'CREATE TABLE IF NOT EXISTS ' + cars_table_name +
        '('
        '    model_year YEAR, '
        '    manufacturer VARCHAR(40), '
        '    product VARCHAR(40)'
        ');')
    cur.execute(SQL)
    db.commit()

    # Python list of dictionaries
    # More info at:
    #     https://stackoverflow.com/questions/8653516/python-list-of-dictionaries-search
    car_data = [
      { 'model_year': '2010', 'manufacturer': 'Toyota', 'product': 'Prius' },
      { 'model_year': '2010', 'manufacturer': 'Honda', 'product': 'CR-V' },
      { 'model_year': '1998', 'manufacturer': 'Honda', 'product': 'Civic' },
      { 'model_year': '1997', 'manufacturer': 'Ford', 'product': 'F-150' },
      { 'model_year': '2017', 'manufacturer': 'Tesla', 'product': 'Model 3' },
    ]

    # Code adapted from: 
    #     https://dev.mysql.com/doc/connector-python/en/connector-python-example-cursor-transaction.html
    add_cars = ('INSERT INTO ' + cars_table_name + ' (model_year, manufacturer, product) '
                '    VALUES (%(model_year)s, %(manufacturer)s, %(product)s)')

    for car_datum in car_data:
        cur.execute(add_cars, car_datum);
    db.commit()

    # Create manufacturer data -----------------------------------------------
    manufacturer_table_name = 'manufacturer'
    SQL = (
        'CREATE TABLE IF NOT EXISTS ' + manufacturer_table_name +
        '('
        '    name VARCHAR(40), '
        '    headquarters VARCHAR(40), '
        '    number_of_employees INT, '
        '    website VARCHAR(40)'
        ');')
    cur.execute(SQL)
    db.commit()

    add_manufacturer = (
        'INSERT INTO ' + manufacturer_table_name + 
        ' (name, headquarters, number_of_employees, website) '
        '    VALUES (%s, %s, %s, %s)')

    # Python list of lists
    # More info at:
    #     https://stackoverflow.com/questions/18449360/access-item-in-a-list-of-lists
    # Data from:
    # https://en.wikipedia.org/wiki/Toyota
    # Honda data from: https://en.wikipedia.org/wiki/Honda
    # Ford data from: https://en.wikipedia.org/wiki/Ford
    # Tesla data from: https://en.wikipedia.org/wiki/Tesla,_Inc.
    manufacture_data = [
      [ 'Toyota', 'Toyota, Aichi, Japan', '364445', 'http://toyota-global.com/' ],
      [ 'Honda', 'Minato, Tokyo, Japan', '208399', 'http://world.honda.com/' ],
      [ 'Ford', 'Dearborn, Michigan, U.S.', '201000', 'http://www.ford.com/' ],
      [ 'Tesla, Inc.', 'Palo Alto, California, US', '33000', 'http://www.tesla.com/' ],
    ]

    for manufacturer_datum in manufacture_data:
        cur.execute(add_manufacturer, manufacturer_datum);
    db.commit()

    # Create Excel (.xlsx) file -----------------------------------------------
    wb = Workbook()

    SQL = 'SELECT * from '+ cars_table_name + ';'
    cur.execute(SQL)
    results = cur.fetchall()
    ws = wb.create_sheet(0)
    ws.title = cars_table_name
    ws.append(cur.column_names)
    for row in results:
        ws.append(row)

    SQL = 'SELECT * from '+ manufacturer_table_name + ';'
    cur.execute(SQL)
    results = cur.fetchall()
    ws = wb.create_sheet(0)
    ws.title = manufacturer_table_name
    ws.append(cur.column_names)
    for row in results:
        ws.append(row)

    workbook_name = "test_workbook"
    wb.save(workbook_name + ".xlsx")

    # Remove tables and database ----------------------------------------------
    SQL = 'DROP TABLE ' + manufacturer_table_name + ';'
    cur.execute(SQL)
    db.commit()
    SQL = 'DROP TABLE ' + cars_table_name + ';'
    cur.execute(SQL)
    db.commit()
    SQL = 'DROP DATABASE ' + database + ';'
    cur.execute(SQL)
    db.commit()

if  __name__ =='__main__':main() 